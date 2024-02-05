# -*- coding: utf-8 -*-
""" Модуль утилит для анализа таблиц графиков """

import os
from datetime import datetime
from typing import Union, List, Tuple
import re

import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from log_settings import log
from models.db import Users, UrlGoogleSheets_month, GroupId
from config_reader import config


file_path = os.path.join(os.path.abspath('files'), f"data.xlsx")
# ключ сервисного аккаунта гугл
credentials_file_path = config.google_credentials_file_path

async def get_today_employees(file_path: Union[str, bool], day: str) -> str:
    """
    Открывает таблицу .xlsx по адресу, и возвращает строку со списком работников:
    [Определяет строку с числами календаря] ->  [находит столбец с переданной датой] ->
    -> [Проходится по 1-му столбцу и заносит всех у кого стоят часы(цифры) в столбце с определенной датой]
    День принимается двухзначным числом (str).
    :param file_path: адрес файла
    :param day: день, например - 01
    """
    # если при загрузке таблицы за нужный месяц не оказалось ссылки
    if file_path is False:
        return "Ссылка на таблицу за этот месяц не прикреплена"

    wb = openpyxl.load_workbook(file_path)                  # Открытие файла Excel
    sheet = wb.active                                       # Выбор активного листа
    chosen_day_column: Union[int, None] = None              # Столбец с выбранной датой

    # поиск столбца с сегодняшней датой - today_date_column
    for row in range(1, sheet.max_row + 1):
        days_list: List[Tuple[int, int]] = []
        flag: bool = False
        # проход по строке, сбор всех двухзначных чисел с номерами колонок в список [days_list]
        for col in range(1, sheet.max_column + 1):
            text_cell = sheet.cell(row=row, column=col).value
            if text_cell and isinstance(text_cell, str):
                # шаблон для числа 01-31
                pattern = r'\b(?:0[1-9]|[12]\d|3[01])\b'
                day_from_column = ''.join(re.findall(pattern, text_cell))
                if day_from_column:
                    days_list.append((col, int(day_from_column)))
        # если количество собранных чисел 30 или 31
        if len(days_list) in [30, 31]:
            for col, d in days_list:
                if int(day) == d:
                    chosen_day_column = col   # номер колонки с сегодняшним числом
                    # log.critical(chosen_day_column)
                    print(f'day input={int(day)}, day from column={int(d)}')
                    flag = True
                    break
        if flag: break
    if chosen_day_column is None:
        log.critical(f"In sheets row not today date ({day})")
        return f"На листе нет информации за указанный день 📆 <code>({day})</code>"

    # Печатаем сотрудников, которые сегодня выходят
    work_list = str()
    work_list += f'<b>{day if day else today_date} числа должны выйти:</b>\n'
    # print("Сегодня выходят:")
    for row in range(1, sheet.max_row + 1):
        employee_name = sheet.cell(row=row, column=1).value
        if not employee_name:
            continue

        # выявляет заголовки по капс локу и обрабатывает их
        if employee_name and any(word.isupper() for word in employee_name.split()) or employee_name.isupper():
            words = re.split(r'[^\w]+', employee_name)                          # сплит по знакам препинания
            employee_name = ' '.join(w for w in words if w.isupper() or w.isdigit())    # джойн только цифр и капсов
            work_list += '\n' + '<b>' + employee_name + ':' + '</b>' + '\n'
            continue

        # добавляет сотрудника, напротив которого цифра в сегодняшней дате
        cell = sheet.cell(row=row, column=chosen_day_column).value
        if cell and str(cell).isdigit():
            work_list += f'• {employee_name}\n'
    # Закрываем файл Excel
    wb.close()
    return work_list


async def save_google_sheet_locally(google_sheets_url: str = None, month: str = None, session: AsyncSession = None,
                              credentials=credentials_file_path) -> Union[str, bool]:
    """ Сохранение гугл таблицы локально. Если не задан url, берет из бд по аргументу 'month'.
    Месяц 'month' принимает двухзначным числом - '01' """
    try:
        # поиск ссылки по месяцу
        if not google_sheets_url:
            query = await session.execute(select(UrlGoogleSheets_month).filter_by(month=month))
            row = query.scalar()
            print(row)

            if row:
                log.warning(f"url for download google sheets - {row.url_google_sheets}")
                if row.url_google_sheets:
                    google_sheets_url = row.url_google_sheets
                elif row.path_file:
                    return row.path_file
            else:
                return False

        # открытие ссылки
        spreadsheet_id = google_sheets_url.split('/')[-2]
        # Устанавливаем разрешения на сервисы:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials, scope)
        gc = gspread.authorize(credentials)
        # Получаем доступ к таблице
        # worksheet = gc.open_by_url(google_sheets_url).sheet1
        worksheet = gc.open_by_key(spreadsheet_id).sheet1
        # Получаем все значения из таблицы
        data = worksheet.get_all_values()

        # сохранение локально
        wb = openpyxl.Workbook()
        ws = wb.active
        # Записываем данные в новую таблицу, преобразуя текст в числа
        for row in data:
            row_as_numbers = [float(cell) if cell.replace('.', '', 1).isdigit() else cell for cell in row]
            ws.append(row_as_numbers)
        # Сохраняем таблицу локально
        wb.save(file_path)
        wb.close()
        log.warning(f'save table locally {file_path}')
        return file_path

    except Exception as exc:
        log.exception(Exception)
        return False


